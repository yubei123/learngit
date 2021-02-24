# encoding: utf-8
from openpyxl import load_workbook, Workbook
from sys import argv
from collections import defaultdict
from datetime import datetime

now = datetime.now().strftime('%Y-%m-%d')

sample = defaultdict()
posP = defaultdict(list)
negP = defaultdict(list)

ps = set()
alls = set()
# 判读表格
wb = load_workbook(argv[1])
ws = wb[wb.sheetnames[0]]
ws.delete_rows(0)
for i in ws.rows:
    name, pathogen, focus = i
    sn = name.value.split('-')[0]
    sample[sn] = 0
    alls.add(sn)
    if focus.value == '-':
        negP[sn].append([name.value, pathogen.value, focus.value])
    elif focus.value.find('疑似') == -1:
        posP[sn].append([name.value, pathogen.value, focus.value])
        ps.add(sn)
    else:
        negP[sn].append([name.value, pathogen.value, focus.value])

ns = alls - ps

outbook = Workbook()

# 病人信息总表
outbook.create_sheet(title='BASIC', index=0)
os = outbook['BASIC']
os.append(['序号', '快递单号', '报告编号', '样本编号*', '医院编号*', '姓名*', '性别*', '年龄*', '住院号*', '床号*', '电话', '送检单位*', \
    '送检科室*', '送检医生*', '采样日期*', '收样日期*', '检测日期*', '报告日期', '报告类型*', '样本类型*', \
    '样本体积*', '样本剩余情况', '主诉', '临床诊断', '临床高度关注病原*', '临床用药*', '是否已使用抗感染药物*', \
    '白细胞计数(WBC)', '淋巴细胞计数', 'C反应蛋白(CRP)', '降钙素原(PCT) ', '中性粒细胞计数', '血小板', '培养结果', '鉴定结果', '镜检结果', '备注', '文件位置'])

wb = load_workbook(argv[2])
ws = wb[wb.sheetnames[0]]

rinfo1 = defaultdict(list)
rinfo2 = defaultdict(list)
report_id = defaultdict(str)

ws.delete_rows(0)
for i in ws.rows:
    if i[3].value in sample:
        b2 = []
        report_id[i[3].value] = i[2].value
        for n, j in enumerate(i):
            if n == 17:
                b2.append(now)
            else:
                b2.append(j.value)
        os.append(b2)
        rinfo1[i[3].value] = ['', '', i[2].value, i[19].value, i[5].value, i[8].value, i[11].value, i[12].value, i[13].value]
        rinfo2[i[3].value] = [i[22].value, i[24].value, i[25].value, i[26].value, i[33].value, i[34].value, i[35].value, i[36].value]

# 报告模版
outbook.create_sheet(title='TPL', index=1)
os = outbook['TPL']
os.append(['sample_id', '结果', '检测类型', 'LOGO', '是否收费', 'tpl_name'])

wb = load_workbook(argv[3])
ws = wb[wb.sheetnames[0]]
ws.delete_rows(0)
boaosample = []
for i in ws.rows:
    name = i[3].value
    if name in sample:
        tpl = []
        tpl.append(name)
        pn = ''
        if name in posP:
            tpl.append('阳性')
            pn = 'positive'
        else:
            tpl.append('阴性')
            pn = 'negative'
        content = []
        if i[4].value:
            if i[4].value.find('收费') > -1:
                content.append('DNA')
            else:
                content.append('免费DNA')
        if i[5].value:
            if i[5].value.find('收费') > -1:
                content.append('RNA')
            else:
                content.append('免费RNA')
        if len(content) == 1:
            tpl.append(content[0])
        else:
            tpl.append('+'.join(content))
        temp = ''
        if i[6].value.find('阿吉安') > -1:
            tpl.append('阿吉安')
            if i[6].value.find('免测') == -1:
                temp = f'0201.mngs.aja.pay.docx'
            else:
                temp = f'0201.mngs.aja.free.docx'
        elif i[6].value.find('郑大一') > -1:
            tpl.append('郑大一')
            if i[6].value.find('盖章') > -1:
                if i[6].value.find('免测') > -1:
                    temp = f'0201.mngs.zju.free.docx'
                else:
                    temp = f'0201.mngs.zju.pay.docx'
            else:
                temp = f'0201.mngs.zju.free.{pn}2.docx'
        elif i[6].value.find('南二医') > -1:
            tpl.append('南二医')
            temp = f'0201.nj2h.docx'
        elif i[6].value.find('华银') > -1:
            tpl.append('华银')
            temp = f'0201.hy.docx'
        elif i[6].value.find('儿童医院') > -1:
            tpl.append('儿童医院')
            temp = f'0201.fzch.docx'
        elif i[6].value.find('明德') > -1:
            tpl.append('明德')
            temp = f'0201.mz.docx'
        elif i[6].value.find('BO') > -1:
            tpl.append('博奥')
            boaosample.append(i[3].value)
            temp = f'0201.boao.docx'
        tpl.append('-')
        tpl.append(temp)
        os.append(tpl)

# 病原汇总表
pathogen = defaultdict(list)
ktype = {'细菌':'bacteria', '真菌':'fungi', '病毒':'virus', '寄生虫':'parasite', '衣原体':'special', '古细菌': 'bacteria', '支原体':'special', '分枝杆菌': 'special'}
wb = load_workbook('D:\病原汇总表.xlsx')
ws = wb[wb.sheetnames[0]]
ws.delete_rows(0)
for i in ws.rows:
    if i[1].value:
        x = [j.value for j in i]
        if x[0]:
            pathogen[x[1]] = [x[2], x[3], x[4], ktype[x[0]], x[5], x[6], x[7], x[0]]
        else:
            pathogen[x[1]] = [x[2], x[3], x[4], '-', x[5], x[6], x[7], '-']

species = defaultdict(list)
def formatS(string):
    if string.find('|') > -1:
        taxid = string.split('|')[0]
        string = string.split('|')[1]
        species[string] = taxid
        return string, 'nt'
    else:
        taxid = string.split('_')[1]
        string = ' '.join(string.split('_')[2:])
        species[string] = taxid
    return string, 'kk'

outbook.create_sheet(title='POS', index=2)
posheet = outbook['POS']
outbook.create_sheet(title='NEG', index=3)
nosheet = outbook['NEG']
outbook.create_sheet(title='Report', index=4)
rs = outbook['Report']
outbook.create_sheet(title='Stat', index=5)
ss = outbook['Stat']

ss.append(['文库编号', '原始reads', '人源reads', '非人源reads', '微生物reads', 'Q30'])
rs.append(['报告日期', '位置', '样本编号', '样本类型', '姓名', '住院号','送检单位','送检科室', '送检医生', \
    '类型', '病原体(genus)', 'reads数目', '病原体(species cn)', '病原体(species en)','reads数目','关注度', \
    '主诉', '临床高度关注病原*', '临床用药*', '是否已使用抗感染药物*', '培养结果', '鉴定结果', '镜检结果', '备注', 'record'])
posheet.append(['sample_id','library_id','kingdom','species_en','species_count','genus_count','abundance','focus',\
    'species_zn','genus_zn','genus_en','type','description','reference'])
nosheet.append(['sample_id','kingdom_zn','species_en','species_zn','species_count','note','description','reference'])

indir = argv[4]

# 读取runstat统计信息
with open(f'{indir}/runStat.xls') as SS:
    SS.readline()
    for i in SS:
        i = i.split('\t')
        ss.append([i[0], int(i[9].replace(',','').replace('"','')),int(i[10].replace(',','').replace('"','')),\
            int(i[12].replace(',','').replace('"','')),int(i[21].replace(',','').replace('"','')),float(i[16])])

# 读取相对丰度表格
for k, v in posP.items():
    for i in v:
        pos = []
        pos.append(k)
        pr = []
        _name, _pathogen, _focus = i
        pos.append(_name)
        sen, ra = formatS(_pathogen)
        if sen in pathogen or sen == 'Pegivirus C' or sen == 'Mycobacteroides abscessus':
            if sen == 'Pegivirus C':
                scn,gen,gcn,kd,t,d,r,_ = pathogen['Human pegivirus']
                pos.extend([kd, sen])
            elif sen == 'Mycobacteroides abscessus':
                scn,gen,gcn,kd,t,d,r,_ = pathogen['Mycobacterium abscessus']
                pos.extend([kd, 'Mycobacterium abscessus'])
            else:
                scn,gen,gcn,kd,t,d,r,_ = pathogen[sen] if sen != 'Pegivirus C' else pathogen['Human pegivirus']
                pos.extend([kd, sen])
            with open(f'{indir}/ra/{_name}.{ra}.ra.xls', encoding='gbk') as rafile:
                rafile.readline()
                for i in rafile:
                    item = i.strip().split('\t')
                    t = t if t else '-'
                    if ra == 'nt':
                        if item[2] == _pathogen:
                            pos.extend([int(item[4]),int(item[5]),float(item[6])])
                            if _focus.strip() in ['高', '低']:
                                pr.extend([t, f'{gcn}{gen}', item[5], scn, sen, item[4], _focus])
                            break
                    else:
                        if item[3] == _pathogen:
                            pos.extend([int(item[5]),int(item[6]),float(item[7])])
                            if _focus.strip() in ['高', '低']:
                                pr.extend([t, f'{gcn}{gen}', item[6], scn, sen, item[5], _focus])
                            break
            pos.extend([_focus, scn, gcn, gen, t, d, r.strip()])
            posheet.append(pos)
            record = f'{ra};RNA' if _name.find('-R') > -1 else f'{ra};DNA'
            rs.append([*rinfo1[k], *pr, *rinfo2[k], record])
        else:
            print(f'{sen} 病原汇总表里面的名字与给出的名字不一致！')

for k, v in negP.items():
    if k in ns:
        rs.append([*rinfo1[k], '阴性', '阴性', '阴性', '阴性', '阴性', '阴性', '阴性', *rinfo2[k]])
    if k in boaosample:
        with open(f'{indir}/../{report_id[k]}_背景列表.xls','w') as boaoback:
            boaoback.write(f'txid\tname\tChinese\thit_reads\n')
            for i in v:
                neg = []
                _name, _pathogen, _focus = i
                sen, ra = formatS(_pathogen)
                if sen in pathogen:
                    scn,gen,gcn,kd,t,d,r,_ = pathogen[sen]
                    neg.extend([species[sen], sen, scn])
                    with open(f'{indir}/ra/{_name}.{ra}.ra.xls', encoding='gbk') as rafile:                    
                        rafile.readline()
                        for i in rafile:
                            item = i.strip().split('\t')
                            if ra == 'nt':
                                if item[2] == _pathogen:
                                    neg.append(str(int(item[4])))
                                    break
                            else:
                                if item[3] == _pathogen:
                                    neg.append(str(int(item[5])))
                                    break
                backstr = '\t'.join(neg)
                boaoback.write(f'{backstr}\n')
        boaoback.close()
    else:
        for i in v:
            neg = []
            neg.append(k)
            _name, _pathogen, _focus = i
            sen, ra = formatS(_pathogen)
            if sen in pathogen:
                scn,gen,gcn,kd,t,d,r,_ = pathogen[sen]
                neg.extend([_, sen, scn])
                with open(f'{indir}/ra/{_name}.{ra}.ra.xls', encoding='gbk') as rafile:
                    rafile.readline()
                    for i in rafile:
                        item = i.strip().split('\t')
                        if ra == 'nt':
                            if item[2] == _pathogen:
                                neg.append(int(item[4]))
                                break
                        else:
                            if item[3] == _pathogen:                                    
                                neg.append(int(item[5]))
                                break
                neg.extend([_focus, d, r.strip()])
                nosheet.append(neg)
            else:
                print(f'{sen} 病原汇总表里面的名字与给出的名字不一致！')                

outbook.save(argv[5])
